from openpyxl import load_workbook
from datetime import datetime
import subprocess
import openpyxl
import shutil
import json
import re
import os

USER = 'tvanderdonck'       # ftp login
PASSWD = 'eZN7Rqd85L9q7h'   # ftp passwd
SERVER = '13.81.52.18'      # ftp host
URI = './automations/a2/downloads/'


from openpyxl import load_workbook
import shutil

def file_treatment(max_nb, file_list):
    commercial_template_path = f'{URI}commercial_template.xlsx'
    fundraising_template_path = f'{URI}fundraising_template.xlsx'
    name = 1

    for file in file_list:
        source_file = f'{URI}{max_nb}/{file}'
        origin_wb = load_workbook(source_file, data_only=True)
        source_sheet = origin_wb['BA Payments New']
        c4_value = source_sheet['C4'].value

        template_wb = load_workbook(fundraising_template_path if c4_value == 'Fundraising' else commercial_template_path)
        target_sheet = template_wb['Feuil1']

        # Copier chaque cellule non fusionnée
        for row in source_sheet.iter_rows():
            for cell in row:
                # Ignorer les cellules fusionnées
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                
                # Copier la valeur de la cellule si elle n'est pas fusionnée
                target_sheet[cell.coordinate].value = cell.value

        # Sauvegarder le fichier modifié
        template_wb.save(fundraising_template_path if c4_value == 'Fundraising' else commercial_template_path)
        
        template_wb.close()
        origin_wb.close()

        # Copier le fichier avec un nom unique
        shutil.copyfile(
            fundraising_template_path if c4_value == 'Fundraising' else commercial_template_path,
            f'{URI}{c4_value}-{str(name)}.xlsx'
        )

        name += 1


def execute(command):  # function to execute command with subprocess
    try:
        result = subprocess.run(
            command, 
            shell=True,
            capture_output=True,
            text=True,
            timeout=10  # Ajout d'un timeout pour éviter les blocages
        )
        return result.stdout

    except subprocess.TimeoutExpired:
        print("La commande a pris trop de temps et a été arrêtée.")
        return "Error"


# *** Partie 1 - Récupérer les xlsx sur le serveur ftp ***
def ftp_mirror():
    print("ftp miror")
    current_year = datetime.now().year  # set current year 
    current_date = f"{datetime.now().year}{datetime.now().month:02}{datetime.now().day:02}"  # set current date
    year_folder_ls_command = f'lftp -u {USER},{PASSWD} ftps://{SERVER} -e "set ssl:verify-certificate no; cd {current_year}; ls; bye"'
    year_folder_ls = execute(year_folder_ls_command)  # exécuter la commande

    # Traiter la sortie et récupérer le dossier le plus récent
    folder_by_date = [int(date) for date in re.findall(r'\b\d{8}\b', year_folder_ls)]
    global last_date_folder
    last_date_folder = max(folder_by_date)

    global new_data_available
    new_data_available = False  # flag initialized

    # Vérifier la date du dernier dossier traité
    try:
        with open('db.json', 'r') as file:
            data = json.load(file)
            if int(data.get("lastdate", 0)) != last_date_folder:
                new_data_available = True
                day_folder_entry_command = f'lftp -u {USER},{PASSWD} ftps://{SERVER} -e "set ssl:verify-certificate no; cd {current_year}; lcd {URI}; mirror {last_date_folder}; ls; bye"'
                execute(day_folder_entry_command)

                # Sauvegarder la nouvelle date dans le fichier JSON
                with open('db.json', 'w') as file_write:
                    json.dump({"lastdate": last_date_folder}, file_write, indent=6)
            else:
                print("Téléchargement déjà à jour.")
    except (json.JSONDecodeError, FileNotFoundError):
        # Créer le fichier si inexistant ou s'il est mal formaté
        with open('db.json', 'w') as file_write:
            json.dump({"lastdate": last_date_folder}, file_write, indent=6)

    return last_date_folder, new_data_available


def run():
    print("run")
    ftp_mirror()
    file_treatment(last_date_folder, os.listdir(f'{URI}{last_date_folder}')) if new_data_available else None
