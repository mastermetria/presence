import openpyxl
import os
import shutil



def file_treatment(max_nb, file_list) :

    commercial_template_path = 'upload/commercial_template.xlsx'
    fundraising_template_path = 'upload/fundraising_template.xlsx'
    name = 1

    for file in file_list:

        source_file = f'upload/ftp_downloads/{max_nb}/{file}'
        origin_wb = openpyxl.load_workbook(source_file, data_only=True)  # 'data_only=True' pour copier uniquement les valeurs
        source_sheet = origin_wb['BA Payments New']  # Remplacer par le nom de la feuille source
        c4_value = source_sheet['C4'].value

        template_wb = openpyxl.load_workbook(fundraising_template_path if c4_value == 'Fundraising' else commercial_template_path)  # Charger le fichier existant


        target_sheet = template_wb['Feuil1']

        for row in source_sheet.iter_rows():
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):

                    continue

                target_sheet[cell.coordinate].value = cell.value

        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))


        template_wb.save(fundraising_template_path if c4_value == 'Fundraising' else commercial_template_path)
        
        template_wb.close()
        origin_wb.close()

        shutil.copyfile(fundraising_template_path if c4_value == 'Fundraising' else commercial_template_path, f'upload/output/{c4_value}-{str(name)}.xlsx')

        name +=1
